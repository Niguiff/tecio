from models import db, Sabor, Insumo, Producto, Venta, ComboItem, Usuario, CierreCaja
from datetime import datetime
from sqlalchemy import extract, func, desc
import pandas as pd
import io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class HeladeriaManager:
    # --- CONSULTAS (READ) ---
    def obtener_sabores_venta(self):
        # Solo activos para el vendedor
        return Sabor.query.filter_by(activo=True).order_by(Sabor.nombre.asc()).all()

    def obtener_todos_sabores(self):
        # Todos (incluso ocultos) para el admin
        return Sabor.query.order_by(Sabor.nombre.asc()).all()

    def obtener_productos(self):
        return Producto.query.all()
    
    def obtener_insumos(self):
        return Insumo.query.all()

    def obtener_ventas(self):
        return Venta.query.all()

    def obtener_usuarios(self):
        return Usuario.query.all()

    # --- STOCK SABORES ---
    def reponer_stock_sabor(self, nombre_sabor, cantidad_gramos, sucursal_destino):
        sabor = Sabor.query.filter_by(nombre=nombre_sabor).first()
        if not sabor: return False, "Sabor no encontrado"

        if sucursal_destino == "Máximo Paz":
            sabor.stock_maximo += cantidad_gramos
        elif sucursal_destino == "Tristán Suárez":
            sabor.stock_tristan += cantidad_gramos
        else:
            return False, "Sucursal desconocida"

        db.session.commit()
        return True, f"Sabor repuesto en {sucursal_destino}."

    def corregir_stock_manual(self, nombre_sabor, baldes_reales, sucursal_destino):
        sabor = Sabor.query.filter_by(nombre=nombre_sabor).first()
        if not sabor: return False, "Sabor no encontrado"
        gramos_reales = baldes_reales * 6000

        if sucursal_destino == "Máximo Paz":
            sabor.stock_maximo = gramos_reales
        elif sucursal_destino == "Tristán Suárez":
            sabor.stock_tristan = gramos_reales
        
        db.session.commit()
        return True, f"Corrección aplicada en {sucursal_destino}."

    # --- STOCK INSUMOS ---
    def reponer_stock_insumo(self, id_insumo, cantidad_unidades, sucursal_destino):
        insumo = Insumo.query.get(id_insumo)
        if not insumo: return False, "Insumo no encontrado"

        if sucursal_destino == "Máximo Paz":
            insumo.stock_maximo += cantidad_unidades
            msg = f"Stock MP: {insumo.stock_maximo}"
        elif sucursal_destino == "Tristán Suárez":
            insumo.stock_tristan += cantidad_unidades
            msg = f"Stock TS: {insumo.stock_tristan}"
        else:
            return False, "Sucursal desconocida"

        db.session.commit()
        return True, f"Insumo repuesto en {sucursal_destino}. {msg}"

    def actualizar_precio(self, id_producto, nuevo_precio):
        prod = Producto.query.get(id_producto)
        if prod:
            prod.precio = nuevo_precio
            db.session.commit()
            return True, "Precio actualizado."
        return False, "Producto no encontrado"

    # --- NUEVA LÓGICA DE TURNOS (CIERRE MANUAL) ---
    def obtener_ventas_turno_actual(self, sucursal):
        """
        Devuelve las ventas realizadas DESDE el último cierre de caja hasta AHORA.
        Si nunca hubo cierre, devuelve todas.
        """
        ultimo_cierre = CierreCaja.query.filter_by(sucursal=sucursal)\
                                        .order_by(CierreCaja.fecha_cierre.desc())\
                                        .first()
        
        query = Venta.query.filter_by(sucursal=sucursal)
        
        if ultimo_cierre:
            # Traer solo ventas posteriores al último cierre
            query = query.filter(Venta.fecha > ultimo_cierre.fecha_cierre)
            
        return query.all()

    def cerrar_caja_sucursal(self, sucursal):
        """
        Realiza el corte: Guarda el registro y 'reinicia' visualmente el contador
        al establecer una nueva marca de tiempo.
        """
        # 1. Calculamos qué estamos cerrando
        ventas_pendientes = self.obtener_ventas_turno_actual(sucursal)
        
        if not ventas_pendientes:
            return False, "No hay ventas nuevas para cerrar."

        total_plata = sum(v.total for v in ventas_pendientes)
        total_cantidad = len(ventas_pendientes)

        # 2. Guardamos el Cierre
        nuevo_cierre = CierreCaja(
            sucursal=sucursal,
            fecha_cierre=datetime.now(),
            monto_total=total_plata,
            cantidad_ventas=total_cantidad
        )
        db.session.add(nuevo_cierre)
        db.session.commit()

        return True, f"Caja de {sucursal} cerrada. Se archivaron ${total_plata}."

    # --- CORE VENTA ---
    def procesar_carrito(self, datos_carrito, sucursal="General"):
        items = datos_carrito.get('items', [])
        medio_pago = datos_carrito.get('medio_pago')
        
        if not items: return False, "Carrito vacío."

        total_a_pagar = 0
        descripcion_venta = []

        try:
            for item in items:
                nombre_prod = item['formato']
                sabores_elegidos = item['sabores'] 
                
                producto = Producto.query.filter_by(nombre=nombre_prod).first()
                if not producto: raise Exception(f"Producto {nombre_prod} no existe")

                total_a_pagar += producto.precio
                
                texto_detalle = f"{producto.nombre}"
                
                # Combos: guardar detalle inmutable
                if producto.es_combo:
                    componentes = ComboItem.query.filter_by(promo_id=producto.id).all()
                    nombres_comp = []
                    for comp in componentes:
                        prod_hijo = Producto.query.get(comp.item_id)
                        if prod_hijo:
                            if comp.cantidad > 1:
                                nombres_comp.append(f"{comp.cantidad}x {prod_hijo.nombre}")
                            else:
                                nombres_comp.append(prod_hijo.nombre)
                    if nombres_comp:
                        texto_detalle += f" [{ ' + '.join(nombres_comp) }]"

                if sabores_elegidos:
                    texto_detalle += f" ({', '.join(sabores_elegidos)})"
                else:
                    texto_detalle += " (Sin sabores)" 
                
                descripcion_venta.append(texto_detalle)
                self._descontar_producto_recursivo(producto, sabores_elegidos, sucursal)

            nueva_venta = Venta(
                fecha=datetime.now(),
                total=total_a_pagar,
                medio_pago=medio_pago,
                detalle="; ".join(descripcion_venta),
                sucursal=sucursal
            )
            db.session.add(nueva_venta)
            db.session.commit()
            
            return True, f"Venta OK. Total: ${total_a_pagar}"

        except Exception as e:
            db.session.rollback()
            return False, f"Error: {str(e)}"

    def _descontar_producto_recursivo(self, producto, lista_sabores_elegidos, sucursal):
        if producto.es_combo:
            items_combo = ComboItem.query.filter_by(promo_id=producto.id).all()
            for comp in items_combo:
                hijo = Producto.query.get(comp.item_id)
                for _ in range(comp.cantidad):
                    self._descontar_producto_recursivo(hijo, lista_sabores_elegidos, sucursal)
            return

        if producto.insumo_id:
            insumo = Insumo.query.get(producto.insumo_id)
            if insumo:
                if sucursal == "Máximo Paz":
                    if insumo.stock_maximo > 0: insumo.stock_maximo -= 1
                elif sucursal == "Tristán Suárez":
                    if insumo.stock_tristan > 0: insumo.stock_tristan -= 1

        if producto.es_helado and producto.peso_helado > 0 and lista_sabores_elegidos:
            peso_por_gusto = producto.peso_helado / len(lista_sabores_elegidos)
            for nombre_sabor in lista_sabores_elegidos:
                sabor_obj = Sabor.query.filter_by(nombre=nombre_sabor).first()
                if sabor_obj:
                    if sucursal == "Máximo Paz":
                        sabor_obj.stock_maximo -= peso_por_gusto
                    elif sucursal == "Tristán Suárez":
                        sabor_obj.stock_tristan -= peso_por_gusto

    # --- REPORTE EXCEL MULTI-HOJA ---
    def generar_reporte_excel(self, fecha_inicio, fecha_fin):
        # 1. Obtener todas las ventas del rango
        ventas_totales = Venta.query.filter(Venta.fecha >= fecha_inicio).filter(Venta.fecha <= fecha_fin).order_by(Venta.fecha.desc()).all()
        
        if not ventas_totales: return None

        # 2. Separar por sucursal
        ventas_mp = [v for v in ventas_totales if v.sucursal == "Máximo Paz"]
        ventas_ts = [v for v in ventas_totales if v.sucursal == "Tristán Suárez"]

        # 3. Crear DataFrames de Detalle
        df_global = self._generar_dataframe_detalle(ventas_totales)
        df_mp = self._generar_dataframe_detalle(ventas_mp)
        df_ts = self._generar_dataframe_detalle(ventas_ts)

        # 4. Crear Buffer de Excel
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            
            # --- HOJA 1: DASHBOARD ---
            self._crear_hoja_dashboard(writer, ventas_totales, ventas_mp, ventas_ts)

            # --- HOJA 2: DETALLE GLOBAL ---
            if not df_global.empty:
                df_global.to_excel(writer, sheet_name='🌎 Detalle Global', index=False)
                self._estilar_hoja_detalle(writer.sheets['🌎 Detalle Global'], df_global)

            # --- HOJA 3: DETALLE MÁXIMO ---
            if not df_mp.empty:
                df_mp.to_excel(writer, sheet_name='📍 Máximo Paz', index=False)
                self._estilar_hoja_detalle(writer.sheets['📍 Máximo Paz'], df_mp)
            
            # --- HOJA 4: DETALLE TRISTÁN ---
            if not df_ts.empty:
                df_ts.to_excel(writer, sheet_name='📍 Tristán Suárez', index=False)
                self._estilar_hoja_detalle(writer.sheets['📍 Tristán Suárez'], df_ts)

        output.seek(0)
        return output

    # --- MÉTODOS PRIVADOS AUXILIARES PARA EL REPORTE ---

    def _crear_hoja_dashboard(self, writer, ventas_global, ventas_mp, ventas_ts):
        """Crea la pestaña de resumen visual con emojis y totales"""
        
        def calcular_metricas(lista_ventas):
            total = sum(v.total for v in lista_ventas)
            cantidad = len(lista_ventas)
            
            efvo = [v for v in lista_ventas if v.medio_pago == 'Efectivo']
            tarj = [v for v in lista_ventas if v.medio_pago == 'Tarjeta']
            qr = [v for v in lista_ventas if v.medio_pago == 'MercadoPago']
            
            return {
                'total_monto': total, 'total_cant': cantidad,
                'efvo_monto': sum(v.total for v in efvo), 'efvo_cant': len(efvo),
                'tarj_monto': sum(v.total for v in tarj), 'tarj_cant': len(tarj),
                'qr_monto': sum(v.total for v in qr), 'qr_cant': len(qr),
            }

        m_global = calcular_metricas(ventas_global)
        m_mp = calcular_metricas(ventas_mp)
        m_ts = calcular_metricas(ventas_ts)

        wb = writer.book
        ws = wb.create_sheet("📊 Dashboard", 0) 
        ws.sheet_view.showGridLines = False

        # Estilos
        titulo_font = Font(size=18, bold=True, color="FFFFFF")
        negrita = Font(bold=True)
        fill_azul = PatternFill("solid", fgColor="0d6efd")
        fill_verde = PatternFill("solid", fgColor="198754")
        fill_gris = PatternFill("solid", fgColor="f8f9fa")
        borde = Border(bottom=Side(style='thin'))

        # --- SECCIÓN 1: TOTAL EMPRESA ---
        ws['B2'] = "RESUMEN GLOBAL DE VENTAS 🌎"
        ws['B2'].font = titulo_font
        ws['B2'].fill = fill_azul
        ws.merge_cells('B2:E2')
        
        data_rows = [
            ("💰 Total Recaudado", f"$ {m_global['total_monto']:,}", f"{m_global['total_cant']} ventas"),
            ("💵 Efectivo", f"$ {m_global['efvo_monto']:,}", f"{m_global['efvo_cant']} ventas"),
            ("💳 Tarjeta", f"$ {m_global['tarj_monto']:,}", f"{m_global['tarj_cant']} ventas"),
            ("📱 QR / MP", f"$ {m_global['qr_monto']:,}", f"{m_global['qr_cant']} ventas"),
        ]
        
        row = 4
        for label, monto, cant in data_rows:
            ws[f'B{row}'] = label
            ws[f'C{row}'] = monto
            ws[f'D{row}'] = cant
            ws[f'B{row}'].font = negrita
            ws[f'B{row}'].border = borde
            row += 1

        # --- SECCIÓN 2: COMPARATIVA POR SUCURSAL ---
        ws['B10'] = "DESGLOSE POR SUCURSAL 🏢"
        ws['B10'].font = titulo_font
        ws['B10'].fill = fill_verde
        ws.merge_cells('B10:E10')

        headers = ["Concepto", "📍 Máximo Paz", "📍 Tristán Suárez"]
        for i, h in enumerate(headers):
            col = get_column_letter(2 + i)
            cell = ws[f'{col}12']
            cell.value = h
            cell.font = negrita
            cell.border = borde
            cell.fill = fill_gris

        comparativa = [
            ("💰 Total ($)", f"$ {m_mp['total_monto']:,}", f"$ {m_ts['total_monto']:,}"),
            ("🛒 Cant. Ventas", m_mp['total_cant'], m_ts['total_cant']),
            ("💵 Efectivo ($)", f"$ {m_mp['efvo_monto']:,}", f"$ {m_ts['efvo_monto']:,}"),
            ("💳 Tarjeta ($)", f"$ {m_mp['tarj_monto']:,}", f"$ {m_ts['tarj_monto']:,}"),
            ("📱 QR ($)", f"$ {m_mp['qr_monto']:,}", f"$ {m_ts['qr_monto']:,}"),
        ]

        row = 13
        for concepto, val_mp, val_ts in comparativa:
            ws[f'B{row}'] = concepto
            ws[f'C{row}'] = val_mp
            ws[f'D{row}'] = val_ts
            row += 1

        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20

    def _generar_dataframe_detalle(self, lista_ventas):
        """Genera el DataFrame detallado para una lista de ventas dada"""
        if not lista_ventas: return pd.DataFrame()

        data_detalle = []
        gran_total = 0
        mapa_precios = {p.nombre: p.precio for p in Producto.query.all()}

        for v in lista_ventas:
            items_texto = v.detalle.split(";")
            for item_raw in items_texto:
                item_raw = item_raw.strip()
                if not item_raw: continue

                nombre_prod = item_raw
                if "(" in item_raw:
                    parts = item_raw.split("(", 1)
                    nombre_prod = parts[0].strip()
                
                fila_item = {
                    "Fecha": v.fecha.strftime("%d/%m/%Y"),
                    "Hora": v.fecha.strftime("%H:%M"),
                    "Sucursal": v.sucursal,
                    "Producto / Items": item_raw,
                    "Medio Pago": v.medio_pago,
                    "Monto ($)": "", 
                    "Tipo Fila": "Item"
                }
                data_detalle.append(fila_item)
            
            fila_subtotal = {
                "Fecha": "", "Hora": "", "Sucursal": "", 
                "Producto / Items": "TOTAL VENTA",
                "Medio Pago": "", "Monto ($)": v.total, "Tipo Fila": "Subtotal"
            }
            data_detalle.append(fila_subtotal)
            gran_total += v.total

        fila_final = {
            "Fecha": "", "Hora": "", "Sucursal": "", 
            "Producto / Items": "TOTAL RECAUDADO",
            "Medio Pago": "", "Monto ($)": gran_total, "Tipo Fila": "GranTotal"
        }
        data_detalle.append(fila_final)
        return pd.DataFrame(data_detalle)

    def _estilar_hoja_detalle(self, ws, df):
        """Aplica colores y bordes a las hojas de detalle"""
        borde = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        fill_header = PatternFill("solid", fgColor="FFC000")
        fill_subtotal = PatternFill("solid", fgColor="E2EFDA")
        fill_grantotal = PatternFill("solid", fgColor="000000")
        font_grantotal = Font(bold=True, color="FFFFFF")
        font_bold = Font(bold=True)

        for cell in ws[1]:
            cell.fill = fill_header
            cell.font = font_bold
            cell.border = borde

        col_tipo_fila = None
        for cell in ws[1]:
            if cell.value == "Tipo Fila":
                col_tipo_fila = cell.column 

        if col_tipo_fila:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                tipo = row[col_tipo_fila-1].value 
                for cell in row:
                    cell.border = borde
                if tipo == "Subtotal":
                    for cell in row:
                        cell.fill = fill_subtotal
                        cell.font = font_bold
                elif tipo == "GranTotal":
                    for cell in row:
                        cell.fill = fill_grantotal
                        cell.font = font_grantotal

            ws.delete_cols(col_tipo_fila)

        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['F'].width = 15